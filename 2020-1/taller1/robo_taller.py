from pyomo.environ import *
import pandas as pd
from openpyxl import load_workbook
from os import getcwd


xlFile = getcwd() + r"\DATOS.XLSX"            
xlFileOut = getcwd() + r"\RESULTADOS.xlsx"            

# Leer información de volumen y peso máximo del camion
book1 = load_workbook(xlFile)
sh = book1["datos2"]

PESOMAX = sh['B1'].value*1000    #Kilogramos

xl_robo = pd.ExcelFile(xlFile)

#LECTURA DE DATOS DE LA BOVEDA;
DATOS_BOVEDA = xl_robo.parse('datos1').set_index(['ARTICULO']) 

#Definir modelo matemático
modeloRobo = ConcreteModel()

#Definir variables de decision
modeloRobo.x = Var( DATOS_BOVEDA.index, domain=NonNegativeReals )

#Definir funcion objetivo
def valorRobo_rule(modeloRobo):

    expr  = sum (DATOS_BOVEDA.VALOR[i]*modeloRobo.x[i] for i in  DATOS_BOVEDA.index) 
    return expr

modeloRobo.Monto_Total = Objective(rule=valorRobo_rule, sense=maximize)

#Definir restricciones
#restriccion de Peso
def rPeso_rule(modeloRobo):
    ecuacion1 = sum(DATOS_BOVEDA.PESO[i]*modeloRobo.x[i] for i in  DATOS_BOVEDA.index)
    return ecuacion1 <= PESOMAX

modeloRobo.r1_Peso = Constraint(rule=rPeso_rule)    

#restriccion de Volumen

#restriccion de Cantidad
def rCantidad_rule(modeloRobo,i):
    ecuacion3 = modeloRobo.x[i]
    return ecuacion3 <= DATOS_BOVEDA.CANTIDAD[i]

modeloRobo.r3_Cantidad = Constraint(DATOS_BOVEDA.index,rule=rCantidad_rule)    

opt = SolverFactory('cbc')

opt_success = opt.solve(modeloRobo)

modeloRobo.write("archivo.lp",io_options={"symbolic_solver_labels":True})

#Imprimir Resultados
print ()
print ("Monto total del robo: ", value(modeloRobo.Monto_Total))
print ()

#Escribir resultados en excel
book = load_workbook(xlFileOut)
writer = pd.ExcelWriter(xlFileOut, engine='openpyxl') 
writer.book = book

out_camion = pd.DataFrame(columns=["artículo","Cantidad"])

fila = 0
for i in DATOS_BOVEDA.index:
    fila += 1
    salida = []
    salida.append(i)
    salida.append(modeloRobo.x[i].value) 
    out_camion.loc[fila] = salida

writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
out_camion.to_excel(writer,'resultado',index=False)

writer.save()
